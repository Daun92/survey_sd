"""
BRIS 컴플레인참조 데이터 API 래퍼
=============================================

⚠️ 실행 환경: 사내 Docker (BRIS 사내망 접근 필요). Vercel 환경 X.
⚠️ 본 파일의 진실원: survey_sd/apps/cs-master/bris_api.py
            (옛 cs/target/bris_api.py 는 archive — 편집 금지)
⚠️ FastAPI 게이트웨이 형태로 노출하려면 lib/bris-api/ 사용 (Phase 3 산출물).

BRIS HTML 페이지를 스크래핑하여 JSON API처럼 사용할 수 있게 해주는 모듈.

사용 방법:
  1) Python 모듈로 import하여 사용
  2) 독립 실행 시 Flask 기반 로컬 API 서버 구동
  3) CSV/Excel 내보내기
  4) Notion 데이터베이스 동기화

예시:
  from bris_api import BrisClient
  client = BrisClient(cookies={"ASP.NET_SessionId": "abc123..."})
  data = client.get_complain_data("2026-04-01", "2026-04-30")

  # Excel 내보내기
  client.to_excel(data, "output.xlsx")

  # CSV 내보내기
  client.to_csv(data, "output.csv")
"""

from bs4 import BeautifulSoup
import re
import json
import csv
import os
import sys
from datetime import datetime, timedelta
from typing import Optional
from pathlib import Path

try:
    import requests
except ImportError:
    requests = None


# ============================================================
# 1. HTML 파서 — lib/bris-parser-py 패키지에서 re-export
# ============================================================
# 파서 코드는 lib/bris-parser-py/src/bris_parser/integrated.py 에 단일 소스 관리.
# 본 모듈은 기존 `from bris_api import BrisParser` 호환을 위해 재노출만 함.

_pkg_src = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'lib', 'bris-parser-py', 'src')
if _pkg_src not in sys.path:
    sys.path.insert(0, _pkg_src)

from bris_parser import BrisParser  # noqa: E402,F401  (재노출)


class BrisSessionExpired(Exception):
    """BRIS 세션이 만료되어 재로그인이 필요한 상태.

    상위 호출자(`bris_to_supabase.py:sync()` 등)가 catch 하여
    ID/PW 재로그인 후 1회 retry 하는 데 사용한다. 레거시의 PermissionError 대신
    이 전용 예외로 통로를 분리해 "응답이 왔지만 구조가 빠짐(silent)"을 포함한다.
    """


# ============================================================
# 2. HTTP 클라이언트 (인증 + Fetch + 파싱 + 내보내기)
# ============================================================

class BrisClient:
    """
    BRIS 페이지를 HTTP로 가져와 파싱하는 클라이언트.

    인증 방법 (택 1):
      A) 쿠키 직접 전달:
         client = BrisClient(cookies={"ASP.NET_SessionId": "abc123"})

      B) 로그인 자동 수행:
         client = BrisClient()
         client.login("user_id", "password")

      C) 브라우저 쿠키 파일 사용:
         client = BrisClient.from_cookie_file("cookies.json")

    데이터 조회:
      records = client.get_complain_data("2026-04-01", "2026-04-30")

    내보내기:
      client.to_csv(records, "output.csv")
      client.to_excel(records, "output.xlsx")
      client.to_json(records, "output.json")
    """

    BASE_URL = "https://bris.exc.co.kr/business/complain_reference_list.asp"
    LOGIN_URL = "https://bris.exc.co.kr/login.asp"

    def __init__(self, session: Optional['requests.Session'] = None,
                 cookies: Optional[dict] = None):
        if requests is None:
            raise ImportError("requests 라이브러리가 필요합니다: pip install requests")
        self.session = session or requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        })
        if cookies:
            self.session.cookies.update(cookies)
        self._logged_in = False

    # ----- 인증 -----

    def login(self, user_id: str, password: str) -> bool:
        """
        BRIS 로그인 수행. 성공 시 True 반환.

        Args:
            user_id:  BRIS 사용자 아이디
            password: BRIS 비밀번호

        Returns:
            True: 로그인 성공 (세션 쿠키 자동 저장)
            False: 로그인 실패
        """
        # 먼저 GET으로 로그인 페이지 접근 (ASP 세션 쿠키 획득)
        self.session.get(self.LOGIN_URL, timeout=30)

        payload = {
            "userid": user_id,
            "passwd": password,
        }
        resp = self.session.post(self.LOGIN_URL, data=payload, timeout=30,
                                 allow_redirects=True,
                                 headers={'Referer': self.LOGIN_URL})

        # 로그인 성공 판별:
        # BRIS는 로그인 성공 시 BRIS_SESSION_TOKEN 쿠키를 발급하고
        # 메인 페이지(frameset)로 리다이렉트함
        session_cookies = {c.name: c.value for c in self.session.cookies}

        # BRIS_SESSION_TOKEN 존재 → 확실히 로그인 성공
        if 'BRIS_SESSION_TOKEN' in session_cookies:
            self._logged_in = True
            return True

        # ASP.NET 세션 쿠키 + 로그인 폼 없음
        has_session = any('session' in k.lower() for k in session_cookies)
        if has_session and 'name=userid' not in resp.text.lower():
            self._logged_in = True
            return True

        # 리다이렉트로 login이 아닌 곳으로 이동
        if resp.status_code == 200 and 'login' not in resp.url.lower():
            self._logged_in = True
            return True

        return False

    def save_cookies(self, filepath: str):
        """현재 세션 쿠키를 JSON 파일로 저장 (재사용 가능)"""
        cookies = {c.name: c.value for c in self.session.cookies}
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(cookies, f, ensure_ascii=False, indent=2)

    @classmethod
    def from_cookie_file(cls, filepath: str) -> 'BrisClient':
        """
        저장된 쿠키 파일로 클라이언트 생성.

        사용법:
          1) 브라우저에서 BRIS 로그인
          2) 개발자도구 → Application → Cookies 에서 쿠키 복사
          3) JSON 파일로 저장: {"쿠키이름": "값", ...}
          4) BrisClient.from_cookie_file("cookies.json")
        """
        with open(filepath, 'r', encoding='utf-8') as f:
            cookies = json.load(f)
        return cls(cookies=cookies)

    @classmethod
    def from_env(cls) -> 'BrisClient':
        """
        환경변수에서 인증 정보를 가져와 로그인.

        환경변수:
          BRIS_USER_ID: 사용자 아이디
          BRIS_PASSWORD: 비밀번호
        """
        user_id = os.environ.get('BRIS_USER_ID')
        password = os.environ.get('BRIS_PASSWORD')
        if not user_id or not password:
            raise ValueError("BRIS_USER_ID, BRIS_PASSWORD 환경변수를 설정하세요")
        client = cls()
        if not client.login(user_id, password):
            raise Exception("BRIS 로그인 실패 - 아이디/비밀번호를 확인하세요")
        return client

    # ----- 데이터 조회 -----

    def get_complain_data(self, start_date: str, end_date: str) -> list[dict]:
        """
        날짜 범위로 컴플레인참조 데이터를 조회하여 dict 리스트로 반환.
        Layer 0 raw HTML 보존이 필요하면 get_complain_data_with_raw() 사용.
        """
        _, records = self.get_complain_data_with_raw(start_date, end_date)
        return records

    def get_complain_data_with_raw(self, start_date: str, end_date: str) -> tuple:
        """
        날짜 범위 조회 — (raw_html, parsed_records) 튜플 반환.
        Layer 0 (cs_bris_raw_pages) 적재용.
        """
        params = {"sDate": start_date, "eDate": end_date}
        resp = self.session.get(self.BASE_URL, params=params, timeout=30)
        resp.encoding = 'utf-8'

        if resp.status_code != 200:
            raise Exception(f"BRIS 요청 실패: HTTP {resp.status_code}")

        # 세션 만료 판별 — 결정적 indicator 는 "목록 테이블이 있는지".
        if not re.search(
            r'<table[^>]*class=["\'][^"\']*ntbl_list_c',
            resp.text,
        ):
            hints = []
            if 'login' in resp.url.lower():
                hints.append(f'url={resp.url}')
            if '로그인' in resp.text[:500]:
                hints.append('body contains 로그인')
            if 'location.href' in resp.text[:200]:
                hints.append('body contains location.href redirect')
            raise BrisSessionExpired(
                f'과정 목록 테이블 없음 (len={len(resp.text)}B'
                + (', ' + ', '.join(hints) if hints else '')
                + '). 세션 만료 추정 — 재로그인 필요.'
            )

        return resp.text, BrisParser.parse(resp.text)

    def get_this_month(self) -> list[dict]:
        """이번 달 데이터 조회"""
        today = datetime.now()
        start = today.replace(day=1).strftime("%Y-%m-%d")
        next_month = today.replace(day=28) + timedelta(days=4)
        end = (next_month.replace(day=1) - timedelta(days=1)).strftime("%Y-%m-%d")
        return self.get_complain_data(start, end)

    def get_today(self) -> list[dict]:
        """오늘 날짜의 과정 목록"""
        today = datetime.now().strftime("%Y-%m-%d")
        return self.get_complain_data(today, today)

    def search(self, start_date: str, end_date: str,
               company: Optional[str] = None,
               course: Optional[str] = None,
               manager: Optional[str] = None,
               echo_status: Optional[str] = None) -> list[dict]:
        """
        데이터 조회 후 추가 필터링.

        Args:
            company: 회사명 키워드
            course:  과정명 키워드
            manager: 수주/수행 담당자 이름
            echo_status: 에코 상태 키워드 (예: "에코 제외", "등록")
        """
        records = self.get_complain_data(start_date, end_date)

        if company:
            records = [r for r in records
                       if company in r.get('회사명', '')
                       or company in r.get('사업장명', '')]
        if course:
            records = [r for r in records
                       if course in r.get('과정명', '')
                       or course in r.get('프로그램명', '')]
        if manager:
            records = [r for r in records
                       if manager in r.get('수주_담당자', '')
                       or manager in r.get('수행_담당자', '')]
        if echo_status:
            records = [r for r in records
                       if echo_status in r.get('에코_상태', '')]

        return records

    # ----- 내보내기 -----

    @staticmethod
    def to_csv(records: list[dict], filepath: str,
               columns: Optional[list[str]] = None, encoding: str = 'utf-8-sig'):
        """
        레코드를 CSV 파일로 내보내기.

        Args:
            records:  파싱된 레코드 리스트
            filepath: 저장할 CSV 파일 경로
            columns:  출력할 컬럼 목록 (기본: 전체 컬럼)
            encoding: 파일 인코딩 (기본: utf-8-sig, Excel 호환)
        """
        cols = columns or BrisParser.COLUMNS
        with open(filepath, 'w', newline='', encoding=encoding) as f:
            writer = csv.DictWriter(f, fieldnames=cols, extrasaction='ignore')
            writer.writeheader()
            for rec in records:
                writer.writerow(rec)
        print(f"CSV 저장 완료: {filepath} ({len(records)}건)")

    @staticmethod
    def to_excel(records: list[dict], filepath: str,
                 columns: Optional[list[str]] = None, sheet_name: str = '컴플레인참조'):
        """
        레코드를 Excel 파일로 내보내기.

        Args:
            records:    파싱된 레코드 리스트
            filepath:   저장할 Excel 파일 경로
            columns:    출력할 컬럼 목록 (기본: 전체 컬럼)
            sheet_name: 시트 이름
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

        cols = columns or BrisParser.COLUMNS
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6",
                                  fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # 헤더 쓰기
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 데이터 쓰기
        data_align = Alignment(vertical="center", wrap_text=True)
        for row_idx, rec in enumerate(records, 2):
            for col_idx, col_name in enumerate(cols, 1):
                val = rec.get(col_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = data_align
                cell.border = thin_border

        # 열 너비 자동 조정
        for col_idx, col_name in enumerate(cols, 1):
            max_len = len(col_name) * 2  # 한글은 2배
            for row_idx in range(2, min(len(records) + 2, 50)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or '')
                # 한글은 2바이트로 계산
                cell_len = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, cell_len)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 40)

        # 필터 설정
        ws.auto_filter.ref = ws.dimensions

        # 첫 행 고정
        ws.freeze_panes = 'A2'

        wb.save(filepath)
        print(f"Excel 저장 완료: {filepath} ({len(records)}건)")

    @staticmethod
    def to_json(records: list[dict], filepath: str):
        """레코드를 JSON 파일로 내보내기"""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        print(f"JSON 저장 완료: {filepath} ({len(records)}건)")


# ============================================================
# 3. Notion 동기화
# ============================================================

class BrisNotionSync:
    """
    BRIS 데이터를 Notion 데이터베이스에 동기화.

    Notion API 키가 필요합니다.
    환경변수 NOTION_API_KEY 또는 생성자에서 직접 전달.

    사용법:
      sync = BrisNotionSync(api_key="ntn_...", database_id="abc123")
      sync.sync_records(records)

    또는 데이터베이스 자동 생성:
      sync = BrisNotionSync(api_key="ntn_...")
      db_id = sync.create_database(parent_page_id="page_id_here")
      sync.sync_records(records)
    """

    NOTION_API_URL = "https://api.notion.com/v1"

    # Notion 속성 타입 매핑
    PROPERTY_SCHEMA = {
        'business_id':        {'type': 'rich_text'},
        'project_id':         {'type': 'rich_text'},
        'echo_id':            {'type': 'rich_text'},
        'customer_id':        {'type': 'rich_text'},
        '과정명':              {'type': 'title'},
        '프로그램명':           {'type': 'rich_text'},
        '과정_총매출':          {'type': 'number', 'format': 'number_with_commas'},
        '시작일':              {'type': 'rich_text'},
        '종료일':              {'type': 'rich_text'},
        '대면_비대면':          {'type': 'select'},
        '수주코드':            {'type': 'rich_text'},
        '수주_프로젝트명':      {'type': 'rich_text'},
        '에코_상태':           {'type': 'select'},
        '수주_담당자':          {'type': 'rich_text'},
        '수주팀':              {'type': 'rich_text'},
        '수행_담당자':          {'type': 'rich_text'},
        '수행팀':              {'type': 'rich_text'},
        '회사명':              {'type': 'rich_text'},
        '사업자번호':           {'type': 'rich_text'},
        '고객_담당자':          {'type': 'rich_text'},
        '고객_이메일':          {'type': 'email'},
    }

    def __init__(self, api_key: Optional[str] = None,
                 database_id: Optional[str] = None):
        if requests is None:
            raise ImportError("requests 라이브러리가 필요합니다")
        self.api_key = api_key or os.environ.get('NOTION_API_KEY', '')
        self.database_id = database_id
        if not self.api_key:
            raise ValueError("Notion API 키가 필요합니다 (api_key 또는 NOTION_API_KEY 환경변수)")
        self.headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
            "Notion-Version": "2022-06-28",
        }

    def create_database(self, parent_page_id: str,
                        title: str = "BRIS 컴플레인참조") -> str:
        """
        Notion 페이지 하위에 데이터베이스를 생성.

        Args:
            parent_page_id: 상위 페이지 ID
            title: 데이터베이스 제목

        Returns:
            생성된 database_id
        """
        properties = {}
        for prop_name, prop_config in self.PROPERTY_SCHEMA.items():
            ptype = prop_config['type']
            if ptype == 'title':
                properties[prop_name] = {"title": {}}
            elif ptype == 'rich_text':
                properties[prop_name] = {"rich_text": {}}
            elif ptype == 'number':
                properties[prop_name] = {
                    "number": {"format": prop_config.get('format', 'number')}
                }
            elif ptype == 'select':
                properties[prop_name] = {"select": {}}
            elif ptype == 'email':
                properties[prop_name] = {"email": {}}

        payload = {
            "parent": {"type": "page_id", "page_id": parent_page_id},
            "title": [{"type": "text", "text": {"content": title}}],
            "properties": properties,
        }

        resp = requests.post(
            f"{self.NOTION_API_URL}/databases",
            headers=self.headers,
            json=payload,
            timeout=30,
        )
        resp.raise_for_status()
        self.database_id = resp.json()["id"]
        print(f"Notion DB 생성 완료: {self.database_id}")
        return self.database_id

    def _record_to_notion_properties(self, record: dict) -> dict:
        """BRIS 레코드를 Notion 페이지 properties 형식으로 변환"""
        props = {}

        for prop_name, prop_config in self.PROPERTY_SCHEMA.items():
            value = record.get(prop_name, '')
            if value is None:
                value = ''
            ptype = prop_config['type']

            if ptype == 'title':
                props[prop_name] = {
                    "title": [{"text": {"content": str(value)}}]
                }
            elif ptype == 'rich_text':
                props[prop_name] = {
                    "rich_text": [{"text": {"content": str(value)}}]
                }
            elif ptype == 'number':
                try:
                    props[prop_name] = {"number": int(value) if value else 0}
                except (ValueError, TypeError):
                    props[prop_name] = {"number": 0}
            elif ptype == 'select' and value:
                props[prop_name] = {"select": {"name": str(value)}}
            elif ptype == 'email' and value:
                props[prop_name] = {"email": str(value)}

        return props

    def sync_records(self, records: list[dict],
                     upsert_key: str = 'business_id') -> dict:
        """
        레코드를 Notion 데이터베이스에 동기화.

        Args:
            records:    BRIS 레코드 리스트
            upsert_key: 중복 판단 기준 필드 (기본: business_id)

        Returns:
            {"created": N, "updated": N, "errors": N}
        """
        if not self.database_id:
            raise ValueError("database_id가 설정되지 않았습니다")

        # 기존 페이지 조회 (upsert용)
        existing = self._get_existing_pages(upsert_key)

        stats = {"created": 0, "updated": 0, "errors": 0}

        for rec in records:
            key_value = str(rec.get(upsert_key, ''))
            properties = self._record_to_notion_properties(rec)

            try:
                if key_value in existing:
                    # 업데이트
                    page_id = existing[key_value]
                    resp = requests.patch(
                        f"{self.NOTION_API_URL}/pages/{page_id}",
                        headers=self.headers,
                        json={"properties": properties},
                        timeout=30,
                    )
                    resp.raise_for_status()
                    stats["updated"] += 1
                else:
                    # 새로 생성
                    resp = requests.post(
                        f"{self.NOTION_API_URL}/pages",
                        headers=self.headers,
                        json={
                            "parent": {"database_id": self.database_id},
                            "properties": properties,
                        },
                        timeout=30,
                    )
                    resp.raise_for_status()
                    stats["created"] += 1
            except Exception as e:
                stats["errors"] += 1
                print(f"Notion 동기화 오류 (business_id={key_value}): {e}")

        print(f"Notion 동기화 완료: 생성 {stats['created']}건, "
              f"업데이트 {stats['updated']}건, 오류 {stats['errors']}건")
        return stats

    def _get_existing_pages(self, key_field: str) -> dict:
        """데이터베이스의 기존 페이지를 key_field 값으로 매핑"""
        pages = {}
        has_more = True
        start_cursor = None

        while has_more:
            payload = {"page_size": 100}
            if start_cursor:
                payload["start_cursor"] = start_cursor

            resp = requests.post(
                f"{self.NOTION_API_URL}/databases/{self.database_id}/query",
                headers=self.headers,
                json=payload,
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()

            for page in data.get("results", []):
                props = page.get("properties", {})
                key_prop = props.get(key_field, {})

                # rich_text 타입에서 값 추출
                if "rich_text" in key_prop and key_prop["rich_text"]:
                    val = key_prop["rich_text"][0].get("plain_text", "")
                    if val:
                        pages[val] = page["id"]

            has_more = data.get("has_more", False)
            start_cursor = data.get("next_cursor")

        return pages


# ============================================================
# 4. Flask 기반 로컬 API 서버 (선택)
# ============================================================

def create_app():
    """Flask 앱 생성 - 로컬 API 서버용"""
    try:
        from flask import Flask, request as flask_request, jsonify
    except ImportError:
        raise ImportError("Flask가 필요합니다: pip install flask")

    app = Flask(__name__)

    @app.route('/api/complain', methods=['GET'])
    def get_complain():
        """
        GET /api/complain?sDate=2026-04-01&eDate=2026-04-30
        선택 파라미터: company, course, manager, echo_status
        """
        s = flask_request.args.get('sDate')
        e = flask_request.args.get('eDate')

        if not s or not e:
            return jsonify({"error": "sDate, eDate 파라미터 필요"}), 400

        try:
            client = BrisClient()
            records = client.search(
                s, e,
                company=flask_request.args.get('company'),
                course=flask_request.args.get('course'),
                manager=flask_request.args.get('manager'),
                echo_status=flask_request.args.get('echo_status'),
            )
            return jsonify({
                "count": len(records),
                "sDate": s,
                "eDate": e,
                "data": records
            })
        except Exception as ex:
            return jsonify({"error": str(ex)}), 500

    @app.route('/api/complain/today', methods=['GET'])
    def get_today():
        """오늘 날짜의 과정 목록"""
        today = datetime.now().strftime("%Y-%m-%d")
        try:
            client = BrisClient()
            records = client.get_complain_data(today, today)
            return jsonify({"count": len(records), "date": today, "data": records})
        except Exception as ex:
            return jsonify({"error": str(ex)}), 500

    @app.route('/api/complain/month', methods=['GET'])
    def get_month():
        """이번 달 전체 데이터"""
        try:
            client = BrisClient()
            records = client.get_this_month()
            return jsonify({"count": len(records), "data": records})
        except Exception as ex:
            return jsonify({"error": str(ex)}), 500

    return app


# ============================================================
# 5. 메인 실행
# ============================================================

if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1 and sys.argv[1] == 'server':
        app = create_app()
        print("BRIS API 서버 시작: http://localhost:5000")
        print("  GET /api/complain?sDate=2026-04-01&eDate=2026-04-30")
        print("  GET /api/complain/today")
        print("  GET /api/complain/month")
        app.run(debug=True, port=5000)

    elif len(sys.argv) > 1 and sys.argv[1] == 'test':
        # 로컬 HTML 파일로 파싱 테스트
        html_file = sys.argv[2] if len(sys.argv) > 2 else 'sample.html'
        with open(html_file, 'r', encoding='utf-8') as f:
            html = f.read()
        records = BrisParser.parse(html)
        print(f"총 {len(records)}건 파싱됨")
        print(json.dumps(records[:2], ensure_ascii=False, indent=2))

    else:
        print("""
BRIS API 래퍼 - 사용법
======================

1) Python에서 import:
   from bris_api import BrisClient

   # 쿠키로 인증
   client = BrisClient(cookies={"ASP.NET_SessionId": "..."})

   # 또는 로그인
   client = BrisClient()
   client.login("user_id", "password")

   # 또는 환경변수
   client = BrisClient.from_env()

   # 데이터 조회
   data = client.get_complain_data("2026-04-01", "2026-04-30")
   data = client.get_this_month()
   data = client.search("2026-04-01", "2026-04-30", company="롯데")

   # 내보내기
   client.to_csv(data, "output.csv")
   client.to_excel(data, "output.xlsx")
   client.to_json(data, "output.json")

2) 로컬 API 서버:
   python bris_api.py server

3) 파싱 테스트:
   python bris_api.py test sample.html

4) Notion 동기화:
   from bris_api import BrisClient, BrisNotionSync
   client = BrisClient(cookies={...})
   data = client.get_complain_data("2026-04-01", "2026-04-30")
   sync = BrisNotionSync(api_key="ntn_...", database_id="...")
   sync.sync_records(data)
        """)
